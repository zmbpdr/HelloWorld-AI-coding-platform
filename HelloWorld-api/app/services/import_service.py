"""题库批量导入服务

支持 Excel (.xlsx) 和 CSV 文件导入，提供两阶段流程：
1. 预检查（preview）：解析文件 → 逐行校验 → 返回错误报告 + 有效数据
2. 确认导入（confirm）：收到确认后，事务入库全部有效行

校验规则：
- 必填字段：language_id、title、slug、question_type
- slug 唯一性：文件内不重复 + 数据库不冲突
- language_id 必须存在于 languages 表中
- question_type 必须是有效枚举值
- difficulty 必须是有效枚举值
- test_cases / options 必须为合法 JSON
- knowledge_tags 支持逗号分隔或 JSON 数组格式
"""

import csv
import io
import json
from typing import Any

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook

from app.models.question import Question
from app.models.course import Language
from app.schemas.admin import VALID_QUESTION_TYPES

# ── 常量 ──────────────────────────────────────────────

VALID_DIFFICULTIES = {"beginner", "intermediate", "advanced"}

REQUIRED_FIELDS = ["language_id", "title", "slug", "question_type"]

# Excel/CSV 列名 → 模型字段名映射（用于表头识别）
FIELD_MAP: dict[str, str] = {
    "language_id": "language_id",
    "title": "title",
    "slug": "slug",
    "question_type": "question_type",
    "description": "description",
    "difficulty": "difficulty",
    "content": "content",
    "answer": "answer",
    "explanation": "explanation",
    "test_cases": "test_cases",
    "starter_code": "starter_code",
    "knowledge_tags": "knowledge_tags",
    "options": "options",
    "order": "order",
}


class ImportService:
    """题库批量导入服务"""

    def __init__(self, db: AsyncSession):
        self.db = db

    # ── 公开方法 ──────────────────────────────────────

    async def preview(self, file: UploadFile) -> dict:
        """解析并校验导入文件，返回预检查报告

        Returns:
            {
                "total_rows": int,
                "valid_rows": int,
                "error_rows": int,
                "errors": [{"row": int, "field": str, "message": str}, ...],
                "valid_data": [...] | None   # 仅当无错误时返回
            }
        """
        content = await file.read()
        filename = (file.filename or "").lower()

        # 解析文件
        if filename.endswith(".csv"):
            rows = self._parse_csv(content)
        elif filename.endswith((".xlsx", ".xls")):
            rows = self._parse_xlsx(content)
        else:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="不支持的文件格式，请上传 .xlsx 或 .csv 文件",
            )

        if not rows:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="文件中没有数据行（仅包含表头或为空）",
            )

        # 从数据库加载校验所需数据
        existing_slugs = await self._get_existing_slugs()
        valid_language_ids = await self._get_language_ids()

        # 逐行校验
        errors: list[dict] = []
        valid_rows: list[dict] = []
        seen_slugs: set[str] = set()

        for i, row in enumerate(rows):
            row_num = i + 2  # 用户看到的行号（1-based，+1 跳过表头）
            row_errors: list[dict] = []

            # ── 必填字段检查 ──
            missing = [
                f for f in REQUIRED_FIELDS
                if f not in row or row[f] is None or str(row[f]).strip() == ""
            ]
            for field in missing:
                row_errors.append({
                    "row": row_num, "field": field, "message": "缺少必填字段",
                })

            if row_errors:
                errors.extend(row_errors)
                continue

            # ── language_id 校验 ──
            try:
                lang_id = int(row["language_id"])
                if lang_id not in valid_language_ids:
                    row_errors.append({
                        "row": row_num, "field": "language_id",
                        "message": f"语言 ID {lang_id} 不存在（有效值: {sorted(valid_language_ids)}）",
                    })
            except (ValueError, TypeError):
                row_errors.append({
                    "row": row_num, "field": "language_id",
                    "message": "language_id 必须是整数",
                })

            # ── slug 唯一性校验 ──
            slug = str(row["slug"]).strip()
            if slug in seen_slugs:
                row_errors.append({
                    "row": row_num, "field": "slug",
                    "message": f"slug '{slug}' 在导入文件中重复",
                })
            elif slug in existing_slugs:
                row_errors.append({
                    "row": row_num, "field": "slug",
                    "message": f"slug '{slug}' 已存在于数据库中",
                })
            else:
                seen_slugs.add(slug)

            # ── question_type 校验 ──
            qt = str(row.get("question_type", "")).strip()
            if qt and qt not in VALID_QUESTION_TYPES:
                row_errors.append({
                    "row": row_num, "field": "question_type",
                    "message": f"无效的题目类型 '{qt}'（有效值: {sorted(VALID_QUESTION_TYPES)}）",
                })

            # ── difficulty 校验 ──
            diff = str(row.get("difficulty", "")).strip()
            if diff and diff not in VALID_DIFFICULTIES:
                row_errors.append({
                    "row": row_num, "field": "difficulty",
                    "message": f"无效的难度 '{diff}'（有效值: {sorted(VALID_DIFFICULTIES)}）",
                })

            # ── JSON 字段格式校验 ──
            for json_field in ["test_cases", "options"]:
                val = row.get(json_field)
                if val and str(val).strip():
                    try:
                        parsed = json.loads(str(val))
                        if not isinstance(parsed, list):
                            row_errors.append({
                                "row": row_num, "field": json_field,
                                "message": f"{json_field} 必须是 JSON 数组格式",
                            })
                    except json.JSONDecodeError:
                        row_errors.append({
                            "row": row_num, "field": json_field,
                            "message": f"{json_field} JSON 格式无效",
                        })

            if row_errors:
                errors.extend(row_errors)
                continue

            # ── 构建有效行数据 ──
            valid_rows.append(self._build_row_data(row, lang_id))

        # 构造报告
        error_row_nums = set(e["row"] for e in errors)
        return {
            "total_rows": len(rows),
            "valid_rows": len(valid_rows),
            "error_rows": len(error_row_nums),
            "errors": errors,
            "valid_data": valid_rows if not errors else None,
        }

    async def confirm_import(self, rows: list[dict], admin_id: int) -> dict:
        """确认导入 — 事务批量插入

        Args:
            rows: 已验证的数据行列表
            admin_id: 操作管理员 ID

        Returns:
            {"imported": int, "message": str}
        """
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="没有可导入的数据",
            )

        questions = [Question(**row) for row in rows]
        self.db.add_all(questions)
        await self.db.commit()

        return {
            "imported": len(questions),
            "message": f"成功导入 {len(questions)} 道题目",
        }

    # ── 文件解析 ──────────────────────────────────────

    def _parse_csv(self, content: bytes) -> list[dict]:
        """解析 CSV 内容"""
        # 尝试 UTF-8，失败则尝试 GBK
        for encoding in ("utf-8-sig", "utf-8", "gbk"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="无法识别文件编码，请使用 UTF-8 或 GBK 编码",
            )

        reader = csv.DictReader(io.StringIO(text))
        header = reader.fieldnames or []
        return self._normalize_rows(reader, header)

    def _parse_xlsx(self, content: bytes) -> list[dict]:
        """解析 Excel 内容"""
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Excel 文件中没有活动工作表",
            )

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c else "" for c in next(rows_iter)]
        except StopIteration:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Excel 文件为空",
            )

        rows = [dict(zip(header, [str(c) if c is not None else "" for c in row])) for row in rows_iter]
        wb.close()
        return self._normalize_rows(rows, header)

    def _normalize_rows(self, rows, header: list[str]) -> list[dict]:
        """统一表头名称并去除空行"""
        # 建立表头映射（支持中英文列名）
        col_map: dict[str, str] = {}
        for col in header:
            col_stripped = col.strip()
            if col_stripped in FIELD_MAP:
                col_map[col] = FIELD_MAP[col_stripped]
            # 也支持直接用英文字段名
            elif col_stripped.lower() in FIELD_MAP:
                col_map[col] = FIELD_MAP[col_stripped.lower()]

        normalized = []
        for row in rows:
            mapped = {}
            for orig_key, value in row.items():
                target_key = col_map.get(orig_key, orig_key.strip().lower() if orig_key else "")
                mapped[target_key] = str(value).strip() if value else ""
            # 跳过完全空行
            if any(v for v in mapped.values()):
                normalized.append(mapped)

        return normalized

    # ── 数据转换 ──────────────────────────────────────

    def _build_row_data(self, row: dict, language_id: int) -> dict:
        """将校验通过的原始行数据转换为模型字段"""

        def parse_json_field(val: str) -> Any:
            """解析 JSON 字段，空值返回 None"""
            if not val or not val.strip():
                return None
            return json.loads(val)

        def parse_tags(val: str) -> list[str]:
            """解析知识点标签：支持逗号分隔或 JSON 数组"""
            if not val or not val.strip():
                return []
            val = val.strip()
            if val.startswith("["):
                return json.loads(val)
            return [t.strip() for t in val.split(",") if t.strip()]

        order_raw = row.get("order", "0")
        try:
            order = int(order_raw) if order_raw.strip() else 0
        except (ValueError, TypeError):
            order = 0

        return {
            "language_id": language_id,
            "title": str(row["title"]).strip(),
            "slug": str(row["slug"]).strip(),
            "question_type": str(row.get("question_type", "coding")).strip() or "coding",
            "description": str(row.get("description", "")).strip() or None,
            "difficulty": str(row.get("difficulty", "beginner")).strip() or "beginner",
            "content": str(row.get("content", "")).strip() or None,
            "answer": str(row.get("answer", "")).strip() or None,
            "explanation": str(row.get("explanation", "")).strip() or None,
            "test_cases": parse_json_field(str(row.get("test_cases", ""))),
            "options": parse_json_field(str(row.get("options", ""))),
            "starter_code": str(row.get("starter_code", "")).strip() or None,
            "knowledge_tags": parse_tags(str(row.get("knowledge_tags", ""))),
            "order": order,
            "is_active": True,
        }

    # ── 数据库查询 ────────────────────────────────────

    async def _get_existing_slugs(self) -> set[str]:
        """获取数据库中已有的所有 slug"""
        result = await self.db.execute(select(Question.slug))
        return {row[0] for row in result.all()}

    async def _get_language_ids(self) -> set[int]:
        """获取数据库中所有有效的语言 ID"""
        result = await self.db.execute(select(Language.id))
        return {row[0] for row in result.all()}